import json
from time import perf_counter
from typing import List, Dict, Tuple, Union, Any

from openpyxl import load_workbook

from data_objects import Location, VehicleType, data_globals
from individual import Individual
from main import Runner
from model import run_settings
from settings import Data
from validation import ArcRoute


def get_exact_output_data_from_sheet(row: int) -> str:
    """Gets the output text file data from a certain row on the solve times summary sheet."""
    workbook = load_workbook(filename="Solve Times Summary.xlsx", read_only=True)
    run_data_sheet = workbook["Run Data"]
    return run_data_sheet[f"H{row}"].value


def extract_data_from_output(math_output: str) -> Tuple[Data, Dict[int, List[List[Union[Tuple[int, int], List[int]]]]]]:
    """Takes the mathematical output text and converts it to input data for the metaheuristic."""
    output_lines = math_output.split("\n")
    # Start after the "Input:" line
    current_line = 1

    # Loop through the customers and compile lists of information about them.
    locations: List[str] = ["Depot"]
    demands: List[int] = [0]
    window_starts: List[int] = [0]
    window_ends: List[int] = [24]
    average_unload_time: List[float] = [0.0]
    coords: List[Tuple[float, float]] = [(0, 0)]
    while output_lines[current_line] and output_lines[current_line].count("Customer ") == 1:
        words = output_lines[current_line].split()
        # The fixed pattern of the output means the same words will always be in the same place
        # print([f"{num}: {word}" for num, word in enumerate(words)])
        # Example line: ['0: Customer', '1: 1', '2: has', '3: 5', '4: pallets', '5: demand', '6: and', '7: window',
        # '8: 0-24', '9: at', '10: (-5.625091957,', '11: 77.494351875)', '12: and', '13: average', '14: unload',
        # '15: time', '16: 0.120062083']
        locations.append(words[1])
        demands.append(int(words[3]))
        window = words[8].split("-")
        window_starts.append(int(window[0]))
        window_ends.append(int(window[1]))
        average_unload_time.append(float(words[16]))
        coords.append((float(words[10][1:-1]), float(words[11][:-1])))

        current_line += 1

    # Add the depot return
    locations.append("DepotReturn")
    demands.append(0)
    window_starts.append(0)
    window_ends.append(32)
    average_unload_time.append(0)
    coords.append((0, 0))

    # Generate distances and times from Euclidean distances between the coordinates.
    distances = []
    times = []
    for index, from_loc in enumerate(coords):
        distances.append([])
        times.append([])
        for to_loc in coords:
            from_x, from_y = from_loc
            to_x, to_y = to_loc
            distance = ((from_x - to_x) ** 2 + (from_y - to_y) ** 2) ** 0.5
            distances[index].append(distance)
            times[index].append(distance / 80)

    # Loop through the vehicles to find information about the vehicle types.
    vehicle_name_types: Dict[str, int] = {}
    vehicle_types: List[str] = []
    distance_cost: List[float] = []
    time_cost: List[float] = []
    pallet_capacity: List[int] = []
    available_vehicles: List[int] = []
    while output_lines[current_line] and output_lines[current_line].count("Vehicle ") == 1:
        # print([f"{num}: {word}" for num, word in enumerate(words)])
        # ['0: Vehicle', '1: SP1', '2: is', '3: a', '4: 11', '5: metre', '6: with', '7: capacity', '8: 30,',
        # '9: distance', '10: cost', '11: 0.796243095,', '12: and', '13: time', '14: cost', '15: 10.888817567']
        # Find the vehicle type name
        line = output_lines[current_line]
        after_name = line.index(" with capacity")
        words: List[str] = line[:after_name].split()
        vehicle_type = " ".join(words[4:])
        vehicle_name = words[1]

        # Check if this vehicle type was already found
        if vehicle_types.count(vehicle_type) == 0:
            # Add the type to the list of types
            vehicle_types.append(vehicle_type)
            available_vehicles.append(1)
            # Get the properties of the vehicle
            words = line[after_name + 1:].split()
            pallet_capacity.append(int(words[2][:-1]))
            distance_cost.append(float(words[5][:-1]))
            time_cost.append(float(words[9]))
        else:
            # Increment the number of available vehicles of this type
            index = vehicle_types.index(vehicle_type)
            available_vehicles[index] += 1

        # Add the index of this vehicle to the dict that will later be used to identify the types of vehicles
        vehicle_name_types[vehicle_name] = vehicle_types.index(vehicle_type)

        current_line += 1

    run_data = Data(locations=locations, vehicle_types=vehicle_types, distance_cost=distance_cost, time_cost=time_cost,
                    pallet_capacity=pallet_capacity, available_vehicles=available_vehicles, demand=demands,
                    window_start=window_starts, window_end=window_ends, average_unload_time=average_unload_time,
                    distances=distances, times=times)

    # Skip blank and output header line
    current_line += 2

    # Iterate across solution lines use dictionary structure to compile a list of which vehicles travel where and to 
    # deliver how much, before compiling this into routes
    all_moves: Dict[str, Dict[str, Dict[str, str]]] = {}
    while output_lines[current_line] and output_lines[current_line].count("Vehicle ") == 1:
        words = output_lines[current_line].split()
        # print([f"{num}: {word}" for num, word in enumerate(words)])
        # ['0: Vehicle', '1: SP1', '2: travels', '3: from', '4: Depot', '5: to', '6: 7', '7: to', '8: deliver', '9: 5',
        # '10: pallets.', '11: Expected', '12: unload', '13: start', '14: time', '15: is', '16: 5.084413751']
        vehicle = words[1]
        if not all_moves.get(vehicle):
            all_moves[vehicle] = {}
        from_loc = words[4]
        all_moves[vehicle][from_loc] = {"to": words[6], "load": words[9]}

        current_line += 1

    # Prepare the mathematical solution dict
    math_solution: Dict[int, List[List[Tuple[int, int]]]] = {}
    for type_index, _ in enumerate(vehicle_types):
        # Create the tour list
        math_solution[type_index] = []

    # Figure out what type each vehicle is, then compile the moves into routes.
    for vehicle, vehicle_moves in all_moves.items():
        # Find the type index for this vehicle
        type_index = vehicle_name_types[vehicle]
        route = []
        # Add the moves to the route
        move = vehicle_moves["Depot"]
        while move["to"] != "DepotReturn":
            route.append((int(move["to"]), int(move["load"])))
            move = vehicle_moves[move["to"]]

        # Add the route to the tour
        math_solution[type_index].append(route)

    return run_data, math_solution


def write_data_to_sheet(row: int = None, exact_routes: Dict[int, List[List[List[int]]]] = None,
                        exact_objective: float = None, meta_routes: Dict[int, List[List[List[int]]]] = None,
                        pretty_meta_routes: str = None, meta_time: float = None, meta_objective: float = None,
                        simple_exact_objective: float = None, simple_meta_objective: float = None):
    """Writes metaheuristic output data to the solve times summary sheet."""
    filename = "Solve Times Summary.xlsx"
    workbook = load_workbook(filename=filename)
    run_data_sheet = workbook["Run Data"]

    if exact_routes:
        run_data_sheet[f"J{row}"].value = json.dumps(exact_routes)
    if exact_objective:
        run_data_sheet[f"K{row}"].value = exact_objective
    if meta_routes:
        run_data_sheet[f"N{row}"].value = json.dumps(meta_routes)
    if pretty_meta_routes:
        run_data_sheet[f"O{row}"].value = pretty_meta_routes
    if meta_time:
        run_data_sheet[f"P{row}"].value = meta_time
    if meta_objective:
        run_data_sheet[f"Q{row}"].value = meta_objective
    if simple_exact_objective:
        run_data_sheet[f"U{row}"].value = simple_exact_objective
    if simple_meta_objective:
        run_data_sheet[f"V{row}"].value = simple_meta_objective

    workbook.save(filename)


def apply_verification_settings(run_data: Data):
    """Applies the input data for the validation to the settings."""
    run_settings.set_run_data(run_data)
    run_settings.RUN_CONFIG.allow_unlimited_fleets = False
    run_settings.RUN_CONFIG.consider_rest_time = False
    run_settings.RUN_CONFIG.return_to_depot_before = 32


def evaluate_solution_simply(solution: Dict[int, List[List[Union[Tuple[int, int], List[int]]]]]) -> float:
    """Evaluates a solution purely based on the cost of distance and time spent travelling."""
    # Grab some variables from the settings so they don't need to be repeatedly retrieved
    depot = data_globals.DEPOT

    # Initialise the overall cost value
    cost: float = 0

    # Loop through each tour
    for vehicle_type_index, tour in solution.items():
        vehicle_type = VehicleType(vehicle_type_index)
        # Track the distance and time of routes in this tour
        distance, time = 0.0, 0.0

        # Loop through each route in the tour
        for route in tour:
            # The distance from the first stop to the depot must be added
            previous_stop = depot
            # Loop through each stop in the route
            for stop_index, stop_data in enumerate(route):
                stop = Location(stop_data[0], stop_data[1])

                # Add the distance between this and the previous stop
                distance += previous_stop.travel_distances[stop.data_index]
                time += previous_stop.travel_times[stop.data_index]

                # The distance from the last stop to the depot must be added
                if stop_index == len(route) - 1:
                    distance += stop.travel_distances[depot.data_index]
                    time += stop.travel_times[depot.data_index]

                previous_stop = stop

        cost += (vehicle_type.distance_cost * distance) + (vehicle_type.time_cost * time)

    return cost


# def verify_constraints(solution_to_verify: Dict[int, List[List[Union[Tuple[int, int], List[int]]]]]):
def verify_constraints(solution_to_verify: Individual):
    """Verify that the given solution meets all the mathematical model's constraints."""
    constraints = {
        "ctDemand": True,
        "ctLoad": True,
        "ctReturnEmpty": True,
        "ctDepart": True,
        "ctReturn": True,
        "ctDepartOnce": True,
        "ctNoReturn": True,
        "ctNoDepart": True,
        "ctRestrictService": True,
        "ctTravel": True,
        "ctWindowStart": True,
        "ctWindowEnd": True,
        "ctWindowTravelTime": True,
        "ctTravelTimeGreater": True
    }

    # Prepare the parameters and variables
    # Parameters
    depot: Location = data_globals.DEPOT
    customers: List[Location] = data_globals.ALL_CUSTOMERS
    depot_return = Location(len(customers) + 1)
    locations: List[Location] = [depot, *customers, depot_return]
    vehicle_types: List[VehicleType] = data_globals.ALL_VEHICLE_TYPES

    class Vehicle:
        """Vehicle class so that the objects can be used as indices (due to the __index__ function)."""

        def __init__(self, vehicle_type: VehicleType):
            self.vehicle_type: VehicleType = vehicle_type
            self._index: int = -1
            self.name: str = ""

        def __index__(self):
            return self._index

        @property
        def index(self) -> int:
            return index

        @index.setter
        def index(self, val: int):
            self._index = val
            self.name = f"SP{val}"

    vehicles = [Vehicle(vehicle_type) for vehicle_type in vehicle_types
                for index in range(1, vehicle_type.available_vehicles + 1)]
    for index, vehicle in enumerate(vehicles):
        vehicle.index = index

    # distances = run_settings.RUN_DATA.distances
    # times = run_settings.RUN_DATA.times
    # distance_cost = run_settings.RUN_DATA.distance_cost
    # time_cost = run_settings.RUN_DATA.time_cost
    # vehicle_types = run_settings.RUN_DATA.vehicle_types
    # available_vehicles = run_settings.RUN_DATA.available_vehicles
    # pallet_capacity = run_settings.RUN_DATA.pallet_capacity
    # demand = run_settings.RUN_DATA.demand
    # window_start = run_settings.RUN_DATA.window_start
    # window_end = run_settings.RUN_DATA.window_end
    # average_unload_time = run_settings.RUN_DATA.average_unload_time

    travel_time_const_matrix: List[List[float]] = [
        [max(i.window_end + i.expected_unload_time + i.travel_times[j] - j.window_start, 0)
         for j in locations]
        for i in locations]
    travel_time_const = 32

    # Variables
    travel = [[[False for _ in vehicles] for _ in locations] for _ in locations]
    deliveries = [[0 for _ in vehicles] for _ in locations]
    service_start = [[i.window_start for _ in vehicles] for i in locations]
    travel_time = [[[0.0 for _ in vehicles] for _ in locations] for _ in locations]

    # Set variables according to solution
    previous_vehicles = 0
    for vehicle_type_index, tour in solution_to_verify.routes.items():
        for route_index, route in enumerate(tour):
            vehicle_index = previous_vehicles + route_index

            prev_stop = depot
            total_delivered = 0
            service_start[prev_stop][vehicle_index] = route.departure_time

            for stop_index, stop in enumerate(route.sequence):

                # Account for the travel and delivery matrix
                travel[prev_stop][stop][vehicle_index] = True
                deliveries[stop][vehicle_index] = stop.serviced_demand

                # Account for the time matrices
                service_start[stop][vehicle_index] = service_start[prev_stop][vehicle_index] + \
                                                     (prev_stop.expected_unload_time) + prev_stop.travel_times[stop]
                if service_start[stop][vehicle_index] < stop.window_start:
                    service_start[stop][vehicle_index] = stop.window_start
                if service_start[stop][vehicle_index] > stop.window_end:
                    raise ValueError(
                        f"Service start {service_start[stop][vehicle_index]} at {stop} is after window end.")
                travel_time[prev_stop][stop][vehicle_index] = service_start[stop][vehicle_index] - \
                                                              service_start[prev_stop][vehicle_index]

                prev_stop = stop
                total_delivered += stop.serviced_demand

                if stop_index == len(route.sequence) - 1:
                    travel[stop][locations[-1]][vehicle_index] = True
                    deliveries[depot][vehicle_index] = total_delivered
                    service_start[depot_return][vehicle_index] = service_start[stop][vehicle_index] + \
                                                                 (stop.expected_unload_time) + \
                                                                 stop.travel_times[depot_return]
                    travel_time[stop][depot_return][vehicle_index] = service_start[depot_return][vehicle_index] - \
                                                                     service_start[stop][vehicle_index]

        previous_vehicles += vehicle_types[vehicle_type_index].available_vehicles

    for j in customers:
        constraints["ctDemand"] = sum(deliveries[j][k] for k in vehicles) == j.demand
        if not constraints["ctDemand"]:
            raise ValueError(f"Constraint broken!\n{constraints}")
    for k in vehicles:
        constraints["ctLoad"] = deliveries[depot][k] == sum(deliveries[j][k] for j in customers)
        constraints["ctReturnEmpty"] = deliveries[depot_return][k] == 0
        constraints["ctDepart"] = deliveries[depot][k] <= k.vehicle_type.capacity * sum(
            travel[depot][j][k] for j in customers)
        constraints["ctReturn"] = sum(travel[depot][j][k] for j in customers) - sum(
            travel[i][depot_return][k] for i in customers) == 0
        constraints["ctDepartOnce"] = sum(travel[depot][j][k] for j in customers) <= 1
        constraints["ctNoReturn"] = sum(travel[i][depot][k] for i in locations) == 0
        constraints["ctNoDepart"] = sum(travel[depot_return][j][k] for j in locations) == 0
        for j in customers:
            constraints["ctRestrictService"] = deliveries[j][k] <= k.vehicle_type.capacity * sum(
                travel[i][j][k] for i in locations)
            constraints["ctTravel"] = sum(travel[i][j][k] for i in locations if i != j) - sum(
                travel[j][i][k] for i in locations) == 0
            if not check_all_true(constraints):
                raise ValueError(f"Constraint broken!\n{constraints}")
        for j in locations:
            constraints["ctWindowStart"] = j.window_start <= service_start[j][
                k]  # or sum(travel[i][j][k] for i in locations) == 0
            constraints["ctWindowEnd"] = service_start[j][
                                             k] <= j.window_end  # or sum(travel[i][j][k] for i in locations) == 0
            for i in locations:
                constraints["ctWindowTravelTime"] = service_start[i][k] + (deliveries[i][k] * i.average_unload_time) + \
                                                    i.travel_times[j] <= service_start[j][k] + (
                                                            (1 - travel[i][j][k]) * travel_time_const_matrix[i][j])
                constraints["ctTravelTimeGreater"] = travel_time[i][j][k] >= service_start[j][k] - service_start[i][
                    k] - (travel_time_const * (1 - travel[i][j][k]))
                if not check_all_true(constraints):
                    raise ValueError(f"Constraint broken!\n{constraints}")

    print(f"All constraints met for {solution}!")


def check_all_true(to_check: Dict[Any, bool]) -> bool:
    """Checks that all values in the dictionary are true."""
    return all(value for value in to_check.values())


if __name__ == "__main__":
    """Verify the metaheuristic against all mathematical instances."""
    # start_row = 38  # The row to start on
    # end_row = start_row + 1  # The row to stop before
    # rows_to_validate = range(49, 57)
    rows_to_validate = [53]
    run_metaheuristic = False
    do_simple_evaluation = False
    eval_cells_in_cols = ["J", "N"]
    verify_constraints_met = False

    for row in rows_to_validate:
        print(f"\nVerifying row {row}\n")

        if run_metaheuristic:
            # Extract the exact data and run the metaheuristic to find a solution and evaluate the exact solution
            text_data = get_exact_output_data_from_sheet(row)
            input_data, exact_solution = extract_data_from_output(text_data)
            apply_verification_settings(input_data)

            start_time = perf_counter()
            runner = Runner(5000, -1, use_multiprocessing=False)
            best_solution: Individual = runner.run()
            end_time = perf_counter()

            # print(f"Row {row}: Pretty output:\n{best_solution.pretty_route_output()}")
            print(f"Row {row}: Run time: {end_time - start_time}")

            best_solution_routes = best_solution.routes_to_dict()
            eval_results = Individual.reconstruct_solution(exact_solution)

            write_data_to_sheet(row=row, exact_routes=exact_solution, exact_objective=eval_results.get_penalised_cost(),
                                meta_routes=best_solution.routes_to_dict(),
                                pretty_meta_routes=best_solution.pretty_route_output(), meta_time=end_time - start_time,
                                meta_objective=best_solution.get_penalised_cost(1))
        else:
            # Extract the exact data and solution, as well as the metaheuristic solution.
            text_data = get_exact_output_data_from_sheet(row)
            input_data, exact_solution = extract_data_from_output(text_data)
            apply_verification_settings(input_data)

            best_solution_routes = ArcRoute.load_json_routes("Solve Times Summary.xlsx", "Run Data", f"N{row}")

        if do_simple_evaluation:
            # Evaluate both solutions fairly, purely based on the objective function of the model.
            simple_exact_objective = evaluate_solution_simply(exact_solution)
            simple_meta_objective = evaluate_solution_simply(best_solution_routes)

            write_data_to_sheet(row=row, simple_exact_objective=simple_exact_objective,
                                simple_meta_objective=simple_meta_objective)

        for col in eval_cells_in_cols:

            solution = Individual.reconstruct_solution(
                ArcRoute.load_json_routes("Solve Times Summary.xlsx", "Run Data", f"{col}{row}"))
            print(solution)
            print(solution.pretty_route_output())

            if verify_constraints_met:
                # Verify that the solution meets all the mathematical model constraints
                solution = Individual.reconstruct_solution(
                    ArcRoute.load_json_routes("Solve Times Summary.xlsx", "Run Data", f"{col}{row}"))
                verify_constraints(solution)
